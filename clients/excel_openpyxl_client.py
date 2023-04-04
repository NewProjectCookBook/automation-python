from __future__ import annotations
import sys
import time
from typing import Optional, Any, Union

import allure
import openpyxl
import psutil
import win32api
import win32com.client
import win32con
import win32process
from openpyxl.cell import Cell, ReadOnlyCell
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelVirtualEngine:
    """ Object of opening Excel file in virtual mode
        Args:
            path_to_the_file: path to Excel file.
            read_only: opening file read-only
            pure_data: getting clean data in cells/formulas
    """
    def __init__(self, path_to_the_file: str, read_only: bool = False, pure_data: bool = False) -> None:
        self.path_to_the_file = path_to_the_file
        self._file = openpyxl.load_workbook(filename=path_to_the_file, read_only=read_only, data_only=pure_data)

    def __enter__(self) -> ExcelVirtualEngine:
        return self

    def __exit__(self, exc_type: Any, exc_val: Any, exc_tb: Any) -> None:
        self._file.close()

    @property
    def excel_file(self) -> Workbook:
        """ Getting an Excel file """
        return self._file

    def get_work_sheet(self, work_sheet_name: str) -> Worksheet:
        """ Getting the contents of a sheet by its name """
        return self.excel_file[work_sheet_name]

    def save_excel_file(self, path_to_save: str) -> None:
        """ Saving an Excel file """
        self.excel_file.save(path_to_save)

    def get_values_from_named_range(self, named_range: str) -> tuple[str, str]:
        """ Getting Information on a Named Range """
        nr = self.excel_file.defined_names[named_range].attr_text.split('!')
        lst = nr[0].replace(r"'", '')
        cells = str(nr[1]).replace('$', '')
        return lst, cells

    def get_sum_of_named_range(self, named_range: str) -> int:
        """ Getting the sum of all values in cells within an ID """
        lst, cells = self.get_values_from_named_range(named_range)
        open_lst = self.get_work_sheet(lst)
        values = []
        for row in open_lst[cells]:
            for cell in row:
                if cell.value and not isinstance(cell.value, str):
                    values.append(cell.value)
        return sum(values)

    def write_data_to_cells(self, named_range: dict) -> None:
        """ Write information to cells by named range """
        with ExcelPhysicalEngine(path_to_the_file=self.path_to_the_file) as ph_work_book:
            for key in named_range:
                lst, cells = self.get_values_from_named_range(key)
                open_lst = self.get_work_sheet(lst)
                data_list = ph_work_book.ph_excel_file.Worksheets(lst)
                for row in open_lst[cells]:
                    for value_from_test_data, cell in zip(named_range.get(key), row):
                        if value_from_test_data:
                            data_list.Cells(cell.row, cell.column).Value = value_from_test_data

    def check_values_in_named_range(self, named_range: dict) -> list[str]:
        """ Checking information in cells against a given named range """
        errors = []
        for key in named_range:
            actual_lst, actual_cells = self.get_values_from_named_range(key)
            open_lst = self.get_work_sheet(actual_lst)
            for row in open_lst[actual_cells]:
                for value_from_test_data, cell in zip(named_range.get(key), row):
                    if isinstance(cell.value, str):
                        cell.value = cell.value.replace(r"'", '')
                        if isinstance(value_from_test_data, str):
                            value_from_test_data = value_from_test_data.replace(r"'", '')
                    if cell.value and cell.value != value_from_test_data:
                        if all(isinstance(value, (float, int)) for value in (cell.value, value_from_test_data)):
                            continue
                        errors.append(
                            f'в ячейке {cell} ожидался результат {value_from_test_data}, а получен {cell.value}')
        return errors

    def get_assumptions_from_sheet(self, name_of_sheet: str) -> list[Cell]:
        """ Getting assembly cells from a model on a specific sheet (green cells) """
        current_list = self.get_work_sheet(name_of_sheet)
        assumptions = [cell for row in current_list[1:current_list.max_row] for cell in row if
                       cell.fill and cell.fill.fgColor.value == 'FFCCFFCC']
        return assumptions

    def compare_sheets_in_excel_files(self, path_to_the_file2: str, name_of_sheet_from_file1: str,
                                      name_of_sheet_from_file2: Optional[str] = None) -> list[str]:
        """ Checking sheets from different Excel files for identity
            If sheets of the same name are compared, then specifying name_of_sheet_from_file2 is optional.
        """
        errors = []
        if not name_of_sheet_from_file2:
            name_of_sheet_from_file2 = name_of_sheet_from_file1
        with ExcelVirtualEngine(path_to_the_file=path_to_the_file2, read_only=True) as work_book2:
            lst1 = self.get_work_sheet(name_of_sheet_from_file1)
            lst2 = work_book2.get_work_sheet(name_of_sheet_from_file2)
            for row1, row2 in zip(lst1.rows, lst2.rows):
                for cell1, cell2 in zip(row1, row2):
                    if cell1.value != cell2.value:
                        errors.append(
                            f'Ячейка {cell1} из файла {self.path_to_the_file}, не равна ячейке {cell2} из файла'
                            f' {path_to_the_file2} {cell1.value} != {cell2.value}')
        return errors

    @staticmethod
    def compare_different_cells_from_excel_files(cells_from_sheet1: list[Union[ReadOnlyCell, Cell]],
                                                 cells_from_sheet2: list[Union[ReadOnlyCell, Cell]],
                                                 soft_comp: bool = False) -> list[str]:
        """ Checking information in cells between two arrays of cells objects.
            You can easily pass arrays of objects from one file to the function for comparison.
        """
        errors = []
        for cell1, cell2 in zip(cells_from_sheet1, cells_from_sheet2):
            if cell1.value != cell2.value:
                if soft_comp and all((not cell1.value, not cell2.value)):
                    continue
                errors.append(f'Ячейка {cell1} из первого файла, не равна ячейке {cell2} из второго файла'
                              f' {cell1.value} != {cell2.value}')
        return errors

    def get_cells(self, name_of_sheet: str, start: str = 'A4', end: str = 'P10') -> list[Cell]:
        """ Getting cells from a model on a specific sheet with a given range
            The default range for the illustrative example is A4:P10
        """
        sheet = self.get_work_sheet(name_of_sheet)
        cell_range = sheet[start:end]
        return [cell for row in cell_range for cell in row]

    def find_max_filling_column(self, name_of_sheet: str, min_row: int, max_row: int, min_col: int) -> str:
        """The function finds a cell that does not contain data, and displays the letter
           designation of the column up to this cell
        """
        with allure.step("Находим буквенное наименование крайнего столбца, в котором есть данные"):
            sheet = self.get_work_sheet(name_of_sheet)
            current_col = None
            for col in sheet.iter_cols(min_row=min_row, max_row=max_row, min_col=min_col, max_col=sheet.max_column+1):
                if col[0].value is None:
                    return current_col
                current_col = col[0].column_letter

    def get_pure_cell_values_from_sheet(self, name_of_sheet: str, start: str = 'A4', end: str = 'P10',
                                        value_round: int = 3) -> list[Any]:
        """ Getting cell values from a model on a specific sheet with a given range
            The default range for the illustrative example is A4:P10
        """
        name_of_sheet = self.get_work_sheet(name_of_sheet)
        cell_range = name_of_sheet[start:end]
        return [round(cell.value, value_round) if isinstance(cell.value, float) else cell.value
                for col in cell_range for cell in col]

    def get_text_index_from_excel_file(self, name_of_sheet: str, text: str = 'Debt/EBITDA', r_cells: bool = False) ->\
            Optional[Union[str, Cell]]:
        """ Getting the line number on which the specified text is located.
            Default line number lookup for text Debt/EBITDA
        """
        excel_list = self.get_work_sheet(name_of_sheet)
        dimensions = excel_list[excel_list.dimensions]
        return next(iter([str(number.row) if not r_cells else number
                          for row in dimensions for number in row if number.value == text]), None)

    def get_dimensions_column_letters_of_timeline(self, name_of_sheet: str)\
            -> list[tuple[str, str]]:
        """ Getting an array of letters of the form [letter of the beginning of the timeline,
            letter of the end of the timeline]
        """
        # TODO Добавить адаптивность функции, переделать определение периодов
        excel_list = self.get_work_sheet(name_of_sheet)
        name_of_periods = [period.column_letter for period in excel_list[1]
                           if ''.join(el for el in str(period.value) if el.isnumeric()).isdigit()]
        return name_of_periods[::len(name_of_periods) - 1]

    def get_work_sheets(self) -> list[str]:
        """ Function to get a list of sheet names """
        worksheet_names = self.excel_file.sheetnames
        return worksheet_names

    def change_worksheet_title(self, name_of_sheet: str, new_name: str) -> None:
        """ Function to change sheet name """
        self.get_work_sheet(name_of_sheet).title = new_name
        self.excel_file.save(self.path_to_the_file)

    def get_dictionary_based_values(self, excel_data: dict[str, Union[list, str]]) -> dict[str, Any]:
        """ Function for collecting data from Excel by given keys and getting a dict with the results """
        result = {}
        for name_of_param, params in excel_data.items():
            if not isinstance(params, list):
                params = [params]
            for param in params:
                work_sheet_name, range_of_cells = self.get_values_from_named_range(named_range=param)
                cells = self.get_work_sheet(work_sheet_name=work_sheet_name)[range_of_cells]
                if not isinstance(cells, tuple):
                    if result.get(name_of_param):
                        result[name_of_param].append(cells.value)
                    else:
                        result[name_of_param] = [cells.value]
                else:
                    values = [(self.get_work_sheet(work_sheet_name=work_sheet_name)[f'{cell.column_letter}2'].value,
                               cell.value) for cell in cells[0] if cell.value]
                    if result.get(name_of_param):
                        result[name_of_param].extend(values)
                    else:
                        result |= {name_of_param: values}
        return result


class ExcelPhysicalEngine:
    """ Object of opening Excel file in physical mode """

    def __init__(self, path_to_the_file: str) -> None:
        self._wait_excel_process()
        self.path_to_the_file = path_to_the_file
        self._excel = win32com.client.Dispatch("Excel.Application")
        self._wb = None

    def __enter__(self) -> ExcelPhysicalEngine:
        try:
            with allure.step(f'Работа с процессом {self.get_current_pid()}'):
                self._wb = self._excel.workbooks.open(self.path_to_the_file, None, False)
                self._excel.Visible = False
                self._excel.DisplayAlerts = False
        except Exception: # noqa
            self.__exit__(*sys.exc_info())
        return self

    @property
    def ph_excel_file(self) -> Any:
        """ Getting an Excel file
                Returns:
                    (COM Object): copy of Excel file.
        """
        return self._wb

    def __exit__(self, exc_type: Any, exc_val: Any, exc_tb: Any) -> None:
        try:
            self._wb.Close(True, self.path_to_the_file)
            self._excel.Visible = True
            self._excel.DisplayAlerts = True
            del self._wb
        finally:
            if hasattr(self, '_excel'):
                try:
                    self._excel.Application.Quit()
                    self._excel.Quit()
                    self.forced_destruction_of_process()
                except Exception: # noqa
                    pass  # ignore, see realization of excel engine

    def get_current_pid(self):
        return win32process.GetWindowThreadProcessId(self._excel.Hwnd)[1]

    def forced_destruction_of_process(self) -> None:
        """ Excel file process force termination generated during object lifetime """
        thr, pid = win32process.GetWindowThreadProcessId(self._excel.Hwnd)
        handle = win32api.OpenProcess(win32con.PROCESS_TERMINATE, 0, pid)
        if handle:
            win32api.TerminateProcess(handle, 0)
            win32api.CloseHandle(handle)

    @staticmethod
    def _wait_excel_process() -> None:
        c_time = time.time()

        def wait_timeout(excel_process):
            with allure.step(f"Ожидание завершения процесса с pid {process.pid}"):
                timeout = time.time() + 60 * 3
                while time.time() < timeout and excel_process.is_running():
                    time.sleep(5)

        with allure.step("Проверка открытых процессов excel"):
            while processes := [proc for proc in psutil.process_iter() if 'EXCEL.EXE' in proc.name()]:
                for process in processes:
                    if process.create_time() > c_time - 60 * 5:
                        wait_timeout(excel_process=process)
                    if process.is_running():
                        with allure.step(f'Завершение процесса с pid = {process.pid}'):
                            process.kill()

    def write_data_to_cell_within_excel(self, cells: list[Cell], data: list[Any]) -> None:
        """ Write information to cells. Recording occurs cell by cell.
            If you need to skip any cell, then in data you need to write "" in place of this cell
        """
        for cell, value_from_test_data in zip(cells, data):
            data_list = self.ph_excel_file.Worksheets(cell.parent.title)
            if value_from_test_data:
                data_list.Cells(cell.row, cell.column).Value = value_from_test_data

    def unprotect_sheets(self, lists: list[str], pw_str: str) -> None:
        """ Function to remove password from sheets
                Args:
                    lists: list of sheets to unlock.
                    pw_str: sheet password
            """
        for list_name in lists:
            data_list = self.ph_excel_file.Worksheets(list_name)
            data_list.Unprotect(pw_str)

    def recalculation_of_model(self) -> None:
        self._excel.Iteration = True
